from openpyxl import Workbook

# Create workbook
wb = Workbook()

# -----------------------
# Sheet 1: Projected Annual Sales
# -----------------------
ws1 = wb.active
ws1.title = "Projected_Annual_Sales"

headers = ["Plant Type", "2026", "2027", "2028", "2029", "2030"]
ws1.append(headers)

data_sales = [
    ["Plant A", 1.5, 1.1, 0.8, 0.5, 0.0],
    ["Plant B", 0.0, 0.1, 0.3, 0.5, 0.6],
    ["Plant C", 0.6, 1.2, 0.6, 1.2, 0.6],
]

for row in data_sales:
    ws1.append(row)

# -----------------------
# Sheet 2: Yield Per Shelf
# -----------------------
ws2 = wb.create_sheet(title="Yield_Per_Shelf")

headers = ["Plant Type", "Yield per Shelf (tons/year)"]
ws2.append(headers)

data_yield = [
    ["Plant A", 0.5],
    ["Plant B", 0.3],
    ["Plant C", 0.4],
]

for row in data_yield:
    ws2.append(row)

# -----------------------
# Sheet 3: Personnel Requirements
# -----------------------
ws3 = wb.create_sheet(title="Personnel_Requirements")

headers = ["Year", "Total Production (tons)", "Direct Workers", "Team Leaders", "Technicians"]
ws3.append(headers)

years = [2026, 2027, 2028, 2029, 2030]

# Sales per year
plant_a = [1.5, 1.1, 0.8, 0.5, 0.0]
plant_b = [0.0, 0.1, 0.3, 0.5, 0.6]
plant_c = [0.6, 1.2, 0.6, 1.2, 0.6]

# Assumptions from problem:
# 1 worker per 0.1 tons of annual yield
# 5 total shelves used
# 1 team leader per used shelf
# 1 technician per produced plant type (A, B, C if produced that year)

for i in range(len(years)):
    total_tons = plant_a[i] + plant_b[i] + plant_c[i]
    direct_workers = total_tons / 0.1

    # Shelves assumed fully used
    team_leaders = 5

    # Count technicians by active plant types
    technicians = 0
    if plant_a[i] > 0:
        technicians += 1
    if plant_b[i] > 0:
        technicians += 1
    if plant_c[i] > 0:
        technicians += 1

    ws3.append([
        years[i],
        round(total_tons, 2),
        int(direct_workers),
        team_leaders,
        technicians
    ])

# Save file
wb.save("plant_production_plan.xlsx")
